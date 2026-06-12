#!/usr/bin/env python
from __future__ import annotations

import argparse
import hashlib
import json
import os
import pickle
import re
import sys
import tempfile
import time
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable
from urllib.parse import unquote, urljoin, urlparse

import numpy as np
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pypdf import PdfReader
from requests.utils import requote_uri
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer


ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
RAW = DATA / "raw"
RAW_DOCS = RAW / "documents"
EXTRACTED = DATA / "extracted"
PROCESSED = DATA / "processed"
SUPPLEMENTAL_SOURCES_PATH = ROOT / "supplemental_sources.json"
INDEX_URL = "https://www.ecb.europa.eu/paym/target/target-professional-use-documents-links/t2s/html/index.en.html"
USER_AGENT = "Mozilla/5.0 (compatible; T2SLocalBot/1.0; +local-ingestion)"
CR_PAGE_URL = "https://www.ecb.europa.eu/paym/target/t2s/governance/html/changerequests.en.html"
DOWNLOAD_RETRIES = int(os.environ.get("T2S_DOWNLOAD_RETRIES", "3"))
PROFESSIONAL_USE_PATH_MARKERS = (
    "/paym/target/target-professional-use-documents-links/t2s/",
    "/paym/target/target-professional-use-documents-links/coco/",
    "/paym/target/t2s/profuse/",
    "/paym/target/t2s/governance/html/changerequests",
)
TEXT_EXTENSIONS = {
    ".txt",
    ".xml",
    ".xsd",
    ".csv",
    ".json",
    ".md",
    ".html",
    ".htm",
    ".yaml",
    ".yml",
}
SUPPORTED_ZIP_MEMBER_EXTENSIONS = TEXT_EXTENSIONS | {".pdf", ".xlsx", ".zip"}
MAX_ZIP_MEMBER_BYTES = int(os.environ.get("T2S_ZIP_MEMBER_MAX_BYTES", str(80 * 1024 * 1024)))
MAX_ZIP_MEMBERS_EXTRACTED = int(os.environ.get("T2S_ZIP_MEMBER_MAX_COUNT", "120"))
XLSX_UNIT_MAX_CHARS = int(os.environ.get("T2S_XLSX_UNIT_MAX_CHARS", "1600"))


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def log(message: str) -> None:
    print(f"[t2s-ingest] {message}", flush=True)


def sha1_text(text: str, n: int = 12) -> str:
    return hashlib.sha1(text.encode("utf-8", errors="ignore")).hexdigest()[:n]


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def safe_slug(text: str, max_len: int = 90) -> str:
    text = unquote(text or "")
    text = text.replace("&", " and ")
    text = re.sub(r"[^A-Za-z0-9._ -]+", " ", text)
    text = re.sub(r"\s+", "-", text.strip())
    text = text.strip("-._")
    if not text:
        text = "document"
    return text[:max_len].strip("-._") or "document"


def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"[ \t\r\f\v]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def normalize_url(url: str) -> str:
    return requote_uri(url.strip())


def url_extension(url: str) -> str:
    path = urlparse(url).path
    suffix = Path(unquote(path)).suffix.lower()
    if suffix:
        return suffix
    if "/html/" in path or path.endswith("/"):
        return ".html"
    return ".html"


def previous_header_title(content_box) -> str:
    prev = content_box.previous_sibling
    while prev is not None:
        if getattr(prev, "name", None) == "div" and "header" in (prev.get("class") or []):
            node = prev.select_one(".title")
            return " ".join((node or prev).get_text(" ", strip=True).split())
        prev = prev.previous_sibling
    return ""


def context_for_anchor(anchor, main) -> dict:
    headers: list[str] = []
    for parent in anchor.parents:
        if parent is main:
            break
        if getattr(parent, "name", None) == "div" and "content-box" in (parent.get("class") or []):
            title = previous_header_title(parent)
            if title:
                headers.append(title)
    headers = list(reversed(headers))

    h3_text = ""
    for h3 in anchor.find_all_previous("h3"):
        if h3.find_parent("main") is main:
            h3_text = " ".join(h3.get_text(" ", strip=True).split())
            break

    return {
        "h3": h3_text,
        "accordion": headers,
        "path": [x for x in [h3_text, *headers] if x],
    }


def release_from_text(text: str) -> str:
    text = text or ""
    match = re.search(r"R(20\d{2})[._-]?(NOV|OCT|JUN|MAR)", text, re.I)
    if match:
        return f"R{match.group(1)}.{match.group(2).upper()}"
    match = re.search(r"Release\s+(\d+(?:\.\d+)?)", text, re.I)
    if match:
        return f"Release {match.group(1)}"
    if re.search(r"March\s+2024", text, re.I):
        return "Ad-hoc March 2024"
    return ""


def doc_family(title: str, url: str, context: dict) -> str:
    hay = " ".join([title, url, " ".join(context.get("path", []))]).lower()
    if "production problem" in hay:
        return "production_problems"
    if "pricing" in hay:
        return "pricing"
    if "change request" in hay:
        return "change_requests"
    if "scope defining" in hay or "sdd" in hay:
        return "t2s_sdd"
    if "user requirement" in hay or "t2s urd" in hay:
        return "t2s_urd"
    if "uhb" in hay or "user handbook" in hay:
        return "t2s_uhb"
    if "udfs" in hay or "detailed functional" in hay:
        return "t2s_udfs"
    if "schema" in hay or "xsd" in hay or "iso 20022" in hay or "message" in hay:
        return "messages_and_schemas"
    if "mept" in hay or "esmig" in hay or "message exchange" in hay or "connectivity" in hay or "nsp" in hay or "gosign" in hay:
        return "connectivity"
    if "business process" in hay or "bpd" in hay:
        return "business_processes"
    if "migration" in hay or "readiness" in hay or "testing" in hay:
        return "migration_and_testing"
    if "hosting terms" in hay or "legal" in hay or "guideline" in hay or "framework agreement" in hay:
        return "legal"
    if "training" in hay or "workshop" in hay or "live demo" in hay or "one pager" in hay or "validations" in hay:
        return "training_and_featured_topics"
    if "participation" in hay or "onboarding" in hay or "csd" in hay or "dcp" in hay:
        return "participation"
    if "release" in hay or release_from_text(hay):
        return "release_documentation"
    if "consultative group" in hay or "meetdoc" in hay:
        return "consultative_group"
    if "shared documentation" in hay:
        return "shared_features"
    return "general"


def category_for(title: str, url: str, context: dict) -> str:
    family = doc_family(title, url, context)
    release = release_from_text(" ".join([title, url, " ".join(context.get("path", []))]))
    if release:
        return "releases"
    return family


def revision_status(title: str, url: str) -> str:
    hay = f"{title} {url}".lower()
    if "with revisions" in hay or "_rev" in hay or "-rev" in hay:
        return "with_revisions"
    if "clean" in hay:
        return "clean"
    return ""


@dataclass
class Document:
    id: str
    title: str
    url: str
    ext: str
    category: str
    family: str
    release: str = ""
    revision_status: str = ""
    contexts: list[dict] = field(default_factory=list)
    status: str = "pending"
    local_path: str = ""
    source_host: str = ""
    media_skipped: bool = False
    error: str = ""
    sha256: str = ""
    size_bytes: int = 0
    extracted_chars: int = 0
    extracted_units: int = 0
    text_path: str = ""
    message_id: str = ""
    usage_guideline_name: str = ""
    collection: str = ""
    publishing_date: str = ""
    mystandards_status: str = ""

    def to_dict(self) -> dict:
        return dict(self.__dict__)


def parse_index(html: str, include_media: bool = False, base_url: str = INDEX_URL) -> tuple[list[Document], list[dict]]:
    soup = BeautifulSoup(html, "lxml")
    main = soup.find("main") or soup
    by_url: dict[str, Document] = {}
    skipped: list[dict] = []

    for anchor in main.find_all("a", href=True):
        title = " ".join(anchor.get_text(" ", strip=True).split())
        href = urljoin(base_url, anchor["href"])
        href = normalize_url(href)
        parsed = urlparse(href)
        if not title or href.endswith("#") or parsed.scheme in {"mailto", "javascript"}:
            continue

        host = parsed.netloc.lower()
        ext = url_extension(href)
        is_media = ext in {".mp4", ".mov", ".avi"}
        is_swift_login = "login.swift.com" in host
        if is_swift_login:
            skipped.append({"title": title, "url": href, "reason": "login_only_external_reference"})
            continue
        if is_media and not include_media:
            skipped.append({"title": title, "url": href, "reason": "media_not_text_document"})
            continue

        context = context_for_anchor(anchor, main)
        family = doc_family(title, href, context)
        release = release_from_text(" ".join([title, href, " ".join(context.get("path", []))]))
        category = category_for(title, href, context)
        key = href
        if key not in by_url:
            doc_id = sha1_text(key)
            by_url[key] = Document(
                id=doc_id,
                title=title,
                url=href,
                ext=ext,
                category=category,
                family=family,
                release=release,
                revision_status=revision_status(title, href),
                contexts=[context],
                source_host=host,
                media_skipped=is_media and not include_media,
            )
        else:
            doc = by_url[key]
            if context not in doc.contexts:
                doc.contexts.append(context)
            if not doc.release and release:
                doc.release = release
            if doc.family == "general" and family != "general":
                doc.family = family
            if doc.category == "general" and category != "general":
                doc.category = category

    docs = list(by_url.values())
    docs.sort(key=lambda d: (d.category, d.release, d.family, d.title.lower()))
    return docs, skipped


def is_crawlable_professional_use_page(url: str) -> bool:
    parsed = urlparse(url)
    if parsed.netloc.lower() != "www.ecb.europa.eu":
        return False
    path = parsed.path.lower()
    if not path.endswith(".html"):
        return False
    if any(marker in path for marker in PROFESSIONAL_USE_PATH_MARKERS):
        return True
    return False


def fetch_html_page(session: requests.Session, url: str, cache_name: str, force: bool) -> str:
    RAW.mkdir(parents=True, exist_ok=True)
    target = RAW / cache_name
    if target.exists() and not force:
        return target.read_text(encoding="utf-8", errors="ignore")
    response = session.get(url, timeout=60)
    response.raise_for_status()
    html = response.text
    target.write_text(html, encoding="utf-8")
    return html


def fetch_index(session: requests.Session, force: bool) -> str:
    target = "ecb_t2s_index.html"
    legacy = ROOT / "ecb_t2s_index.raw.html"
    if legacy.exists() and not force:
        html = legacy.read_text(encoding="utf-8", errors="ignore")
        RAW.mkdir(parents=True, exist_ok=True)
        (RAW / target).write_text(html, encoding="utf-8")
        return html
    return fetch_html_page(session, INDEX_URL, target, force)


DEFAULT_MAX_PAGES = int(os.environ.get("T2S_MAX_CRAWL_PAGES", "128"))


def discover_professional_use_docs(
    session: requests.Session,
    *,
    force: bool,
    include_media: bool,
    max_pages: int,
) -> tuple[list[Document], list[dict], list[str]]:
    queued = [INDEX_URL, CR_PAGE_URL]
    visited: set[str] = set()
    by_url: dict[str, Document] = {}
    skipped: list[dict] = []

    while queued and len(visited) < max_pages:
        page_url = queued.pop(0)
        if page_url in visited:
            continue
        visited.add(page_url)
        cache_name = f"ecb_t2s_page_{sha1_text(page_url)}.html"
        html = fetch_html_page(session, page_url, cache_name, force=force)
        docs, page_skipped = parse_index(html, include_media=include_media, base_url=page_url)
        skipped.extend(page_skipped)
        log(f"parsed {len(docs)} links from professional-use page: {page_url}")
        for doc in docs:
            existing = by_url.get(doc.url)
            if existing:
                for context in doc.contexts:
                    if context not in existing.contexts:
                        existing.contexts.append(context)
                if not existing.release and doc.release:
                    existing.release = doc.release
                if existing.family == "general" and doc.family != "general":
                    existing.family = doc.family
                if existing.category == "general" and doc.category != "general":
                    existing.category = doc.category
                continue
            by_url[doc.url] = doc
            if is_crawlable_professional_use_page(doc.url) and doc.url not in visited and doc.url not in queued:
                queued.append(doc.url)
    if queued:
        log(
            f"crawl stopped at max_pages={max_pages} with {len(queued)} professional-use pages still queued; "
            "increase --max-pages or T2S_MAX_CRAWL_PAGES for a wider crawl"
        )

    docs = list(by_url.values())
    docs.sort(key=lambda d: (d.category, d.release, d.family, d.title.lower()))
    return docs, skipped, sorted(visited)


def local_path_for(doc: Document) -> Path:
    release = safe_slug(doc.release, 40) if doc.release else ""
    folder = RAW_DOCS / safe_slug(doc.category, 60)
    if release:
        folder = folder / release
    suffix = doc.ext if doc.ext.startswith(".") else f".{doc.ext}"
    name = safe_slug(doc.title, 80)
    return folder / f"{name}__{doc.id}{suffix}"


def download_doc(session: requests.Session, doc: Document, force: bool = False) -> Document:
    target = local_path_for(doc)
    target.parent.mkdir(parents=True, exist_ok=True)
    doc.local_path = str(target.relative_to(ROOT))

    if target.exists() and target.stat().st_size > 0 and not force:
        doc.status = "downloaded"
        doc.size_bytes = target.stat().st_size
        doc.sha256 = sha256_file(target)
        return doc

    last_error: Exception | None = None
    for attempt in range(1, max(1, DOWNLOAD_RETRIES) + 1):
        try:
            with session.get(doc.url, stream=True, timeout=120, allow_redirects=True) as response:
                response.raise_for_status()
                tmp = target.with_suffix(target.suffix + ".tmp")
                with tmp.open("wb") as fh:
                    for block in response.iter_content(chunk_size=1024 * 256):
                        if block:
                            fh.write(block)
                tmp.replace(target)
            doc.status = "downloaded"
            doc.error = ""
            doc.size_bytes = target.stat().st_size
            doc.sha256 = sha256_file(target)
            return doc
        except Exception as exc:
            last_error = exc
            if attempt < max(1, DOWNLOAD_RETRIES):
                time.sleep(min(2 ** attempt, 10))
    doc.status = "failed"
    doc.error = repr(last_error)
    log(f"download failed: {doc.title} -> {last_error}")
    return doc


def extract_pdf(path: Path) -> list[dict]:
    units: list[dict] = []
    reader = PdfReader(str(path))
    if reader.is_encrypted:
        try:
            reader.decrypt("")
        except Exception:
            pass
    for idx, page in enumerate(reader.pages, start=1):
        try:
            text = clean_text(page.extract_text() or "")
        except Exception as exc:
            text = f"[page extraction failed: {exc!r}]"
        if text:
            units.append({"unit_type": "page", "unit": idx, "text": text})
    return units


def extract_html(path: Path) -> list[dict]:
    raw = path.read_text(encoding="utf-8", errors="ignore")
    soup = BeautifulSoup(raw, "lxml")
    for tag in soup(["script", "style", "noscript", "svg", "header", "footer", "nav"]):
        tag.decompose()
    main = soup.find("main") or soup.body or soup
    title = soup.find("title")
    parts = []
    if title:
        parts.append(" ".join(title.get_text(" ", strip=True).split()))
    for node in main.find_all(["h1", "h2", "h3", "h4", "p", "li", "td", "th"]):
        text = " ".join(node.get_text(" ", strip=True).split())
        if text:
            parts.append(text)
    text = clean_text("\n".join(parts))
    return [{"unit_type": "html", "unit": 1, "text": text}] if text else []


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def extract_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    units: list[dict] = []
    for ws in wb.worksheets:
        header = ""
        block_lines: list[str] = []
        block_chars = 0
        block_no = 1

        def emit_block() -> None:
            nonlocal block_lines, block_chars, block_no
            if not block_lines:
                return
            prefix = f"Sheet: {ws.title}"
            if header:
                prefix += f"\nColumns: {header}"
            text = clean_text(prefix + "\n" + "\n".join(block_lines))
            if text:
                units.append({"unit_type": "sheet", "unit": f"{ws.title} block {block_no}", "text": text})
            block_lines = []
            block_chars = 0
            block_no += 1

        row_no = 0
        for row in ws.iter_rows(values_only=True):
            values = []
            for value in row:
                cell = _cell_text(value)
                if cell:
                    values.append(cell)
            if values:
                row_no += 1
                line = " | ".join(values)
                if not header:
                    header = line
                row_line = f"Row {row_no}: {line}"
                if block_lines and block_chars + len(row_line) + len(header) + 32 > XLSX_UNIT_MAX_CHARS:
                    emit_block()
                block_lines.append(row_line)
                block_chars += len(row_line) + 1
        emit_block()
    wb.close()
    return units


def _extract_path_by_extension(path: Path, suffix: str) -> list[dict]:
    if suffix == ".pdf":
        return extract_pdf(path)
    if suffix in {".html", ".htm"}:
        return extract_html(path)
    if suffix == ".xlsx":
        return extract_xlsx(path)
    if suffix == ".zip":
        return extract_zip(path)
    if suffix in TEXT_EXTENSIONS:
        text = clean_text(path.read_text(encoding="utf-8", errors="ignore"))
        return [{"unit_type": "file", "unit": 1, "text": text}] if text else []
    return []


def extract_zip(path: Path) -> list[dict]:
    units: list[dict] = []
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        inventory = "\n".join(names)
        units.append({"unit_type": "zip_inventory", "unit": 1, "text": f"ZIP inventory:\n{inventory}"})
        extracted_count = 0
        with tempfile.TemporaryDirectory(prefix="t2s_zip_") as tmpdir:
            tmp_root = Path(tmpdir)
            for name in names:
                suffix = Path(name).suffix.lower()
                info = archive.getinfo(name)
                if info.is_dir() or suffix not in SUPPORTED_ZIP_MEMBER_EXTENSIONS:
                    continue
                if info.file_size > MAX_ZIP_MEMBER_BYTES:
                    units.append(
                        {
                            "unit_type": "zip_member_skipped",
                            "unit": name,
                            "text": f"ZIP member skipped because it is too large: {name} ({info.file_size} bytes)",
                        }
                    )
                    continue
                if extracted_count >= MAX_ZIP_MEMBERS_EXTRACTED:
                    units.append(
                        {
                            "unit_type": "zip_member_skipped",
                            "unit": name,
                            "text": f"ZIP member skipped because archive extraction reached {MAX_ZIP_MEMBERS_EXTRACTED} supported members",
                        }
                    )
                    continue
                try:
                    raw = archive.read(name)
                    safe_name = safe_slug(Path(name.replace("\\", "/")).name or "member", 70)
                    member_path = tmp_root / f"{extracted_count:04d}_{safe_name}{suffix}"
                    member_path.write_bytes(raw)
                    inner_units = _extract_path_by_extension(member_path, suffix)
                except Exception as exc:
                    units.append({"unit_type": "zip_member_error", "unit": name, "text": f"ZIP member extraction failed: {name} ({exc!r})"})
                    continue
                extracted_count += 1
                for inner in inner_units:
                    text = inner.get("text")
                    if text:
                        units.append(
                            {
                                "unit_type": f"zip_{inner.get('unit_type') or 'member'}",
                                "unit": f"{name} :: {inner.get('unit')}",
                                "text": f"ZIP member: {name}\n{text}",
                            }
                        )
    return units


def extract_text_for_doc(doc: Document) -> tuple[Document, list[dict]]:
    path = ROOT / doc.local_path
    units: list[dict] = []
    try:
        if doc.ext == ".pdf":
            units = extract_pdf(path)
        elif doc.ext in {".html", ".htm", ".rss"}:
            units = extract_html(path)
        elif doc.ext == ".xlsx":
            units = extract_xlsx(path)
        elif doc.ext == ".zip":
            units = extract_zip(path)
        else:
            try:
                units = [{"unit_type": "file", "unit": 1, "text": clean_text(path.read_text(encoding="utf-8", errors="ignore"))}]
            except UnicodeDecodeError:
                units = []
        text = "\n\n".join(f"[{u['unit_type']} {u['unit']}]\n{u['text']}" for u in units if u.get("text"))
        EXTRACTED.mkdir(parents=True, exist_ok=True)
        text_path = EXTRACTED / f"{doc.id}.txt"
        text_path.write_text(text, encoding="utf-8")
        doc.text_path = str(text_path.relative_to(ROOT))
        doc.extracted_chars = len(text)
        doc.extracted_units = len(units)
        if not units:
            doc.error = (doc.error + " " if doc.error else "") + "no_text_extracted"
    except Exception as exc:
        doc.error = (doc.error + " " if doc.error else "") + f"extract_failed:{exc!r}"
        log(f"extract failed: {doc.title} -> {exc}")
    return doc, units


def split_text(text: str, max_chars: int = 1800, overlap: int = 250) -> list[str]:
    text = clean_text(text)
    if not text:
        return []
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    chunks: list[str] = []
    current = ""
    for para in paragraphs:
        if len(para) > max_chars:
            if current:
                chunks.append(current.strip())
                current = ""
            start = 0
            while start < len(para):
                chunks.append(para[start : start + max_chars].strip())
                start += max_chars - overlap
            continue
        if len(current) + len(para) + 2 <= max_chars:
            current = f"{current}\n\n{para}" if current else para
        else:
            if current:
                chunks.append(current.strip())
            tail = current[-overlap:] if current and overlap else ""
            current = f"{tail}\n\n{para}".strip() if tail else para
    if current:
        chunks.append(current.strip())
    return [c for c in chunks if len(c) > 40]


def build_chunks(docs: list[Document], extracted: dict[str, list[dict]]) -> list[dict]:
    chunks: list[dict] = []
    for doc in docs:
        if doc.status != "downloaded":
            continue
        context_path = []
        for ctx in doc.contexts:
            for item in ctx.get("path", []):
                if item and item not in context_path:
                    context_path.append(item)
        metadata_text = clean_text(
            "\n".join(
                [
                    f"Document title: {doc.title}",
                    f"Document id: {doc.id}",
                    f"Extension: {doc.ext}",
                    f"Category: {doc.category}",
                    f"Family: {doc.family}",
                    f"Release: {doc.release}" if doc.release else "",
                    f"Revision status: {doc.revision_status}" if doc.revision_status else "",
                    f"Publishing date: {doc.publishing_date}" if doc.publishing_date else "",
                    f"Context path: {' > '.join(context_path)}" if context_path else "",
                    f"Local path: {doc.local_path}",
                    f"Source URL: {doc.url}",
                ]
            )
        )
        if metadata_text:
            chunks.append(
                {
                    "chunk_id": f"{doc.id}:document_metadata:0:0",
                    "doc_id": doc.id,
                    "title": doc.title,
                    "category": doc.category,
                    "family": doc.family,
                    "release": doc.release,
                    "revision_status": doc.revision_status,
                    "publishing_date": doc.publishing_date,
                    "context_path": context_path,
                    "unit_type": "document_metadata",
                    "unit": 0,
                    "text": metadata_text,
                    "local_path": doc.local_path,
                    "source_url": doc.url,
                }
            )
        units = extracted.get(doc.id, [])
        for unit in units:
            pieces = split_text(unit.get("text", ""))
            for idx, piece in enumerate(pieces):
                chunk_id = f"{doc.id}:{unit.get('unit_type')}:{unit.get('unit')}:{idx}"
                chunks.append(
                    {
                        "chunk_id": chunk_id,
                        "doc_id": doc.id,
                        "title": doc.title,
                        "category": doc.category,
                        "family": doc.family,
                        "release": doc.release,
                        "revision_status": doc.revision_status,
                        "publishing_date": doc.publishing_date,
                        "context_path": context_path,
                        "unit_type": unit.get("unit_type"),
                        "unit": unit.get("unit"),
                        "text": piece,
                        "local_path": doc.local_path,
                        "source_url": doc.url,
                    }
                )
    return chunks


def build_index(docs: list[Document], chunks: list[dict]) -> None:
    PROCESSED.mkdir(parents=True, exist_ok=True)
    docs_payload = [doc.to_dict() for doc in docs]
    (PROCESSED / "documents.json").write_text(json.dumps(docs_payload, indent=2, ensure_ascii=False), encoding="utf-8")
    with (PROCESSED / "documents.jsonl").open("w", encoding="utf-8") as fh:
        for doc in docs_payload:
            fh.write(json.dumps(doc, ensure_ascii=False) + "\n")
    with (PROCESSED / "chunks.jsonl").open("w", encoding="utf-8") as fh:
        for chunk in chunks:
            fh.write(json.dumps(chunk, ensure_ascii=False) + "\n")

    if not chunks:
        raise RuntimeError("No chunks were produced; cannot build index")

    search_texts = [
        "\n".join(
            [
                chunk.get("title", ""),
                chunk.get("family", ""),
                chunk.get("category", ""),
                chunk.get("release", ""),
                chunk.get("message_id", ""),
                chunk.get("usage_guideline_name", ""),
                chunk.get("collection", ""),
                chunk.get("schema_file", ""),
                chunk.get("schema_component", ""),
                chunk.get("component_name", ""),
                chunk.get("schema_path", ""),
                chunk.get("schema_target_namespace", ""),
                chunk.get("mystandards_status", ""),
                chunk.get("publishing_date", ""),
                chunk.get("unit_type", ""),
                str(chunk.get("unit", "")),
                chunk.get("local_path", ""),
                chunk.get("source_url", ""),
                " > ".join(chunk.get("context_path", [])),
                chunk.get("text", ""),
            ]
        )
        for chunk in chunks
    ]
    word_vectorizer = TfidfVectorizer(
        lowercase=True,
        strip_accents="unicode",
        ngram_range=(1, 3),
        min_df=1,
        max_features=260000,
        sublinear_tf=True,
        token_pattern=r"(?u)\b[\w./-]{2,}\b",
    )
    char_vectorizer = TfidfVectorizer(
        lowercase=True,
        strip_accents="unicode",
        analyzer="char_wb",
        ngram_range=(3, 5),
        min_df=1,
        max_features=180000,
        sublinear_tf=True,
    )
    log(f"building word index over {len(chunks)} chunks")
    word_matrix = word_vectorizer.fit_transform(search_texts)
    log("building character index")
    char_matrix = char_vectorizer.fit_transform(search_texts)
    bm25_vectorizer = CountVectorizer(
        lowercase=True,
        strip_accents="unicode",
        ngram_range=(1, 2),
        min_df=1,
        max_features=220000,
        token_pattern=r"(?u)\b[\w./-]{2,}\b",
    )
    log("building BM25 lexical index")
    bm25_matrix = bm25_vectorizer.fit_transform(search_texts).tocsr()
    bm25_doc_len = np.asarray(bm25_matrix.sum(axis=1)).ravel().astype(float)
    bm25_avgdl = float(bm25_doc_len.mean() or 1.0)
    bm25_df = np.asarray((bm25_matrix > 0).sum(axis=0)).ravel().astype(float)
    bm25_idf = np.log(((len(chunks) - bm25_df + 0.5) / (bm25_df + 0.5)) + 1.0)
    payload = {
        "built_at": utc_now(),
        "index_url": INDEX_URL,
        "docs": docs_payload,
        "chunks": chunks,
        "chunk_id_to_pos": {chunk.get("chunk_id"): idx for idx, chunk in enumerate(chunks) if chunk.get("chunk_id")},
        "word_vectorizer": word_vectorizer,
        "word_matrix": word_matrix,
        "char_vectorizer": char_vectorizer,
        "char_matrix": char_matrix,
        "bm25_vectorizer": bm25_vectorizer,
        "bm25_matrix": bm25_matrix,
        "bm25_idf": bm25_idf,
        "bm25_doc_len": bm25_doc_len,
        "bm25_avgdl": bm25_avgdl,
    }
    with (PROCESSED / "index.pkl").open("wb") as fh:
        pickle.dump(payload, fh, protocol=pickle.HIGHEST_PROTOCOL)

    summary = {
        "built_at": payload["built_at"],
        "source": INDEX_URL,
        "documents_total": len(docs),
        "documents_downloaded": sum(1 for d in docs if d.status == "downloaded"),
        "documents_failed": sum(1 for d in docs if d.status == "failed"),
        "chunks": len(chunks),
        "index_flavour": "premium_hybrid_tfidf_char_bm25",
        "bm25_features": len(bm25_vectorizer.vocabulary_),
        "schema_chunks": sum(1 for chunk in chunks if str(chunk.get("unit_type", "")).startswith("schema")),
        "categories": sorted({d.category for d in docs}),
        "families": sorted({d.family for d in docs}),
    }
    (PROCESSED / "manifest.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")


def rebuild_index_from_processed() -> dict:
    """Rebuild vector indexes from processed JSON files.

    This is used by downstream ingesters, e.g. MyStandards, after appending
    extra local documents/chunks to the processed corpus.
    """
    documents_path = PROCESSED / "documents.json"
    chunks_path = PROCESSED / "chunks.jsonl"
    if not documents_path.exists() or not chunks_path.exists():
        raise FileNotFoundError("Processed documents/chunks not found; run t2s_ingest.py first")
    docs_payload = json.loads(documents_path.read_text(encoding="utf-8"))
    docs = [Document(**doc) for doc in docs_payload]
    chunks: list[dict] = []
    with chunks_path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                chunks.append(json.loads(line))
    build_index(docs, chunks)
    return json.loads((PROCESSED / "manifest.json").read_text(encoding="utf-8"))


def load_supplemental_sources() -> list[dict]:
    if not SUPPLEMENTAL_SOURCES_PATH.exists():
        return []
    data = json.loads(SUPPLEMENTAL_SOURCES_PATH.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError(f"{SUPPLEMENTAL_SOURCES_PATH} must contain a JSON list")
    return [item for item in data if isinstance(item, dict) and item.get("url")]


def supplemental_document(item: dict) -> Document:
    url = normalize_url(str(item["url"]))
    title = str(item.get("title") or Path(urlparse(url).path).stem or url).strip()
    category = str(item.get("category") or "legal").strip() or "legal"
    family = str(item.get("family") or "target_dkk").strip() or "target_dkk"
    release = str(item.get("release") or "").strip()
    publishing_date = str(item.get("publishing_date") or "").strip()
    context_label = str(item.get("context") or "Supplemental authoritative sources").strip()
    return Document(
        id=sha1_text(url),
        title=title,
        url=url,
        ext=url_extension(url),
        category=category,
        family=family,
        release=release,
        contexts=[{"h3": context_label, "accordion": [], "path": [context_label]}],
        source_host=urlparse(url).netloc.lower(),
        publishing_date=publishing_date,
    )


def append_supplemental_sources(force: bool = False) -> dict:
    sources = load_supplemental_sources()
    if not sources:
        log("no supplemental sources configured")
        return rebuild_index_from_processed()
    documents_path = PROCESSED / "documents.json"
    chunks_path = PROCESSED / "chunks.jsonl"
    if not documents_path.exists() or not chunks_path.exists():
        raise FileNotFoundError("Processed documents/chunks not found; run t2s_ingest.py first")

    docs_payload = json.loads(documents_path.read_text(encoding="utf-8"))
    docs = [Document(**doc) for doc in docs_payload]
    chunks: list[dict] = []
    with chunks_path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                chunks.append(json.loads(line))

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT, "Accept": "*/*"})
    supplemental_docs = [supplemental_document(item) for item in sources]
    supplemental_ids = {doc.id for doc in supplemental_docs}
    supplemental_urls = {doc.url for doc in supplemental_docs}

    docs = [doc for doc in docs if doc.id not in supplemental_ids and doc.url not in supplemental_urls]
    chunks = [chunk for chunk in chunks if chunk.get("doc_id") not in supplemental_ids]

    extracted: dict[str, list[dict]] = {}
    downloaded: list[Document] = []
    for idx, doc in enumerate(supplemental_docs, start=1):
        log(f"downloading supplemental {idx}/{len(supplemental_docs)}: {doc.title[:90]}")
        downloaded_doc = download_doc(session, doc, force=force)
        downloaded.append(downloaded_doc)
        if downloaded_doc.status == "downloaded":
            log(f"extracting supplemental {idx}/{len(supplemental_docs)}: {downloaded_doc.title[:90]}")
            extracted_doc, units = extract_text_for_doc(downloaded_doc)
            extracted[extracted_doc.id] = units

    docs.extend(downloaded)
    chunks.extend(build_chunks(downloaded, extracted))
    build_index(docs, chunks)
    skipped_path = PROCESSED / "skipped_links.json"
    skipped = json.loads(skipped_path.read_text(encoding="utf-8")) if skipped_path.exists() else []
    write_catalog(docs, skipped)
    manifest = json.loads((PROCESSED / "manifest.json").read_text(encoding="utf-8"))
    log(f"supplemental done: {len(supplemental_docs)} sources, {manifest['chunks']} chunks")
    return manifest


def load_processed_documents_and_chunks() -> tuple[list[Document], list[dict]]:
    documents_path = PROCESSED / "documents.json"
    chunks_path = PROCESSED / "chunks.jsonl"
    if not documents_path.exists() or not chunks_path.exists():
        raise FileNotFoundError("Processed documents/chunks not found; run t2s_ingest.py first")
    docs_payload = json.loads(documents_path.read_text(encoding="utf-8"))
    docs = [Document(**doc) for doc in docs_payload]
    chunks: list[dict] = []
    with chunks_path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                chunks.append(json.loads(line))
    return docs, chunks


def append_new_current_documents(max_pages: int = DEFAULT_MAX_PAGES, include_media: bool = False) -> dict:
    docs, chunks = load_processed_documents_and_chunks()
    existing_by_url = {doc.url: doc for doc in docs if doc.url}

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT, "Accept": "*/*", "Cache-Control": "no-cache"})
    current_docs, skipped, crawled_pages = discover_professional_use_docs(
        session,
        force=True,
        include_media=include_media,
        max_pages=max_pages,
    )
    new_docs = [doc for doc in current_docs if doc.url not in existing_by_url]
    changed_metadata = 0
    for current in current_docs:
        existing = existing_by_url.get(current.url)
        if not existing:
            continue
        before = (
            existing.title,
            existing.category,
            existing.family,
            existing.release,
            existing.revision_status,
            existing.contexts,
        )
        existing.title = current.title
        existing.category = current.category
        existing.family = current.family
        existing.release = current.release
        existing.revision_status = current.revision_status
        existing.contexts = current.contexts
        after = (
            existing.title,
            existing.category,
            existing.family,
            existing.release,
            existing.revision_status,
            existing.contexts,
        )
        if before != after:
            changed_metadata += 1

    extracted: dict[str, list[dict]] = {}
    downloaded: list[Document] = []
    for idx, doc in enumerate(new_docs, start=1):
        log(f"downloading new document {idx}/{len(new_docs)}: {doc.title[:90]}")
        downloaded_doc = download_doc(session, doc, force=False)
        downloaded.append(downloaded_doc)
        if downloaded_doc.status == "downloaded":
            log(f"extracting new document {idx}/{len(new_docs)}: {downloaded_doc.title[:90]}")
            extracted_doc, units = extract_text_for_doc(downloaded_doc)
            extracted[extracted_doc.id] = units

    docs.extend(downloaded)
    chunks.extend(build_chunks(downloaded, extracted))

    supplemental_sources = load_supplemental_sources()
    supplemental_docs = [supplemental_document(item) for item in supplemental_sources]
    missing_supplemental = [doc for doc in supplemental_docs if doc.url not in {item.url for item in docs}]
    if missing_supplemental:
        log(f"adding {len(missing_supplemental)} missing supplemental sources")
        supplemental_ids = {doc.id for doc in missing_supplemental}
        supplemental_urls = {doc.url for doc in missing_supplemental}
        docs = [doc for doc in docs if doc.id not in supplemental_ids and doc.url not in supplemental_urls]
        chunks = [chunk for chunk in chunks if chunk.get("doc_id") not in supplemental_ids]
        supplemental_extracted: dict[str, list[dict]] = {}
        supplemental_downloaded: list[Document] = []
        for idx, doc in enumerate(missing_supplemental, start=1):
            log(f"downloading supplemental {idx}/{len(missing_supplemental)}: {doc.title[:90]}")
            downloaded_doc = download_doc(session, doc, force=False)
            supplemental_downloaded.append(downloaded_doc)
            if downloaded_doc.status == "downloaded":
                log(f"extracting supplemental {idx}/{len(missing_supplemental)}: {downloaded_doc.title[:90]}")
                extracted_doc, units = extract_text_for_doc(downloaded_doc)
                supplemental_extracted[extracted_doc.id] = units
        docs.extend(supplemental_downloaded)
        chunks.extend(build_chunks(supplemental_downloaded, supplemental_extracted))

    if not new_docs and not missing_supplemental and not changed_metadata:
        log("incremental sync: no new documents or listing metadata changes")
    else:
        log(
            f"incremental sync: new={len(new_docs)}, missing_supplemental={len(missing_supplemental)}, "
            f"metadata_updates={changed_metadata}"
        )
    build_index(docs, chunks)
    write_catalog(docs, skipped)
    (PROCESSED / "crawled_pages.json").write_text(json.dumps(crawled_pages, indent=2, ensure_ascii=False), encoding="utf-8")
    manifest = json.loads((PROCESSED / "manifest.json").read_text(encoding="utf-8"))
    log(f"incremental done: {manifest['documents_downloaded']} docs, {manifest['chunks']} chunks")
    return manifest


def write_catalog(docs: list[Document], skipped: list[dict]) -> None:
    lines = ["# T2S local catalog", "", f"Generated: {utc_now()}", "", "## Documents", ""]
    for doc in sorted(docs, key=lambda d: (d.category, d.release, d.family, d.title.lower())):
        status = doc.status
        release = f" | {doc.release}" if doc.release else ""
        local = doc.local_path or ""
        lines.append(f"- [{doc.category}/{doc.family}{release}] {doc.title} ({status})")
        lines.append(f"  - Local: `{local}`")
        lines.append(f"  - Source: {doc.url}")
    if skipped:
        lines.extend(["", "## Skipped external/media links", ""])
        for item in skipped:
            lines.append(f"- {item.get('title')} - {item.get('reason')} - {item.get('url')}")
    (PROCESSED / "catalog.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    (PROCESSED / "skipped_links.json").write_text(json.dumps(skipped, indent=2, ensure_ascii=False), encoding="utf-8")


def ingest(
    force: bool = False,
    include_media: bool = False,
    limit: int | None = None,
    refresh_index: bool = False,
    max_pages: int = DEFAULT_MAX_PAGES,
) -> dict:
    for folder in [RAW_DOCS, EXTRACTED, PROCESSED]:
        folder.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT, "Accept": "*/*"})
    fetch_index(session, force=force or refresh_index)
    docs, skipped, crawled_pages = discover_professional_use_docs(
        session,
        force=force or refresh_index,
        include_media=include_media,
        max_pages=max_pages,
    )
    if limit:
        docs = docs[:limit]
    log(
        f"found {len(docs)} unique local-document candidates across {len(crawled_pages)} "
        f"professional-use pages; skipped {len(skipped)} external/media links"
    )

    downloaded: list[Document] = []
    for idx, doc in enumerate(docs, start=1):
        log(f"downloading {idx}/{len(docs)}: {doc.title[:90]}")
        downloaded.append(download_doc(session, doc, force=force))
        time.sleep(0.05)

    extracted: dict[str, list[dict]] = {}
    for idx, doc in enumerate(downloaded, start=1):
        if doc.status != "downloaded":
            continue
        log(f"extracting {idx}/{len(downloaded)}: {doc.title[:90]}")
        doc, units = extract_text_for_doc(doc)
        extracted[doc.id] = units

    chunks = build_chunks(downloaded, extracted)
    log(f"built {len(chunks)} retrieval chunks")
    build_index(downloaded, chunks)
    write_catalog(downloaded, skipped)
    (PROCESSED / "crawled_pages.json").write_text(json.dumps(crawled_pages, indent=2, ensure_ascii=False), encoding="utf-8")

    manifest = json.loads((PROCESSED / "manifest.json").read_text(encoding="utf-8"))
    log(f"done: {manifest['documents_downloaded']} docs, {manifest['chunks']} chunks")
    return manifest


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Ingest ECB T2S documents into a local optimized index.")
    parser.add_argument("--force", action="store_true", help="redownload and rebuild from scratch")
    parser.add_argument("--refresh-index", action="store_true", help="fetch the ECB index page again but reuse cached documents when possible")
    parser.add_argument("--include-media", action="store_true", help="also download linked media files such as MP4")
    parser.add_argument("--limit", type=int, default=None, help="debug: ingest only first N documents")
    parser.add_argument("--max-pages", type=int, default=DEFAULT_MAX_PAGES, help="maximum ECB professional-use HTML pages to crawl")
    parser.add_argument("--supplemental-only", action="store_true", help="download configured supplemental sources and rebuild the index")
    parser.add_argument("--incremental", action="store_true", help="discover current documentation and ingest only new URLs before rebuilding vectors")
    args = parser.parse_args(argv)
    try:
        if args.incremental:
            append_new_current_documents(max_pages=args.max_pages, include_media=args.include_media)
        elif args.supplemental_only:
            append_supplemental_sources(force=args.force)
        else:
            ingest(
                force=args.force,
                include_media=args.include_media,
                limit=args.limit,
                refresh_index=args.refresh_index,
                max_pages=args.max_pages,
            )
        return 0
    except KeyboardInterrupt:
        return 130
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

